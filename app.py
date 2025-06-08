from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, jsonify, make_response
import pandas as pd
import os
from werkzeug.utils import secure_filename
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import uuid
from functools import wraps # Import for decorators

app = Flask(__name__)
app.secret_key = 'your-very-secure-and-unguessable-key-12345' # CHANGE THIS IN PRODUCTION!
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PROCESSED_FOLDER'] = 'processed_data'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls', 'csv'}

# Admin Credentials (For demonstration purposes ONLY - use environment variables/config for production!)
app.config['ADMIN_USERNAME'] = 'admin'
app.config['ADMIN_PASSWORD'] = 'password123' # CHANGE THIS IN PRODUCTION!

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)
os.makedirs('static', exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def advanced_matcher(input_val, db_values):
    """Enhanced matching with threshold"""
    input_val_str = str(input_val).lower().replace('_', ' ').replace('-', ' ').strip()
    best_match = None
    best_score = 0

    for db_val in db_values:
        db_val_str = str(db_val)
        db_val_clean = db_val_str.lower().replace('_', ' ').replace('-', ' ').strip()
        score = fuzz.token_sort_ratio(input_val_str, db_val_clean)
        if score > best_score:
            best_score = score
            best_match = db_val # Return original db_val

    return best_match if best_score >= 75 else None

def compare_values(input_val, db_val):
    """Compare values and return formatted result"""
    if pd.isna(input_val) and pd.isna(db_val):
        return "Passed (Both Empty)"
    if pd.isna(input_val) or input_val == 'Not Found' or input_val == '':
        return f"Failed (Input: -, Ref: {str(db_val).strip()})"
    if pd.isna(db_val) or db_val == 'Not Found' or db_val == '':
        return f"Failed (Input: {str(input_val).strip()}, Ref: -)"

    input_str = str(input_val).strip()
    db_str = str(db_val).strip()

    try: # Numeric comparison
        if float(input_str) == float(db_str):
            return "Passed"
    except ValueError: # String comparison
        if input_str.lower() == db_str.lower():
            return "Passed"

    return f"Failed (Input: {input_str}, Ref: {db_str})"

# --- Authentication Decorator ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Please log in to access the admin page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('logged_in'):
        return redirect(url_for('admin')) # Already logged in, redirect to admin

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == app.config['ADMIN_USERNAME'] and password == app.config['ADMIN_PASSWORD']:
            session['logged_in'] = True
            flash('Logged in successfully!', 'success')
            return redirect(url_for('admin'))
        else:
            flash('Invalid username or password.', 'error')
            return render_template('login.html') # Stay on login page with error

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('home'))


@app.route('/admin', methods=['GET', 'POST'])
@login_required # Protect this route with login
def admin():
    if request.method == 'POST':
        if 'database' not in request.files:
            flash('No file part selected', 'error')
            return redirect(request.url)
        file = request.files['database']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            try:
                filename = secure_filename('database_master_file' + os.path.splitext(file.filename)[1])
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                xls = pd.ExcelFile(filepath)
                sheet_names = xls.sheet_names

                session['db_file_path'] = filepath
                session['db_sheet_names'] = sheet_names
                session['db_uploaded'] = True
                session.pop('db_configured', None)
                session.pop('user_input_configured', None)
                session.pop('processed_results_path', None)
                session.pop('result_columns', None)
                session.pop('processing_done', None)

                flash('Database uploaded. Please configure matching parameters.', 'success')
                return redirect(url_for('configure'))
            except Exception as e:
                flash(f'Error processing database file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type. Allowed types are xlsx, xls, csv.', 'error')
            return redirect(request.url)
            
    return render_template('admin.html')

@app.route('/get_sheet_columns/<file_type>', methods=['POST'])
# This endpoint might be called by JS from a protected page, so good to protect it too
@login_required
def get_sheet_columns(file_type):
    sheet_name = request.form.get('sheet_name')
    file_path_session_key = None

    if file_type == 'admin_db':
        file_path_session_key = 'db_file_path'
    elif file_type == 'user_input':
        file_path_session_key = 'user_file_path'
    else:
        return jsonify({'error': 'Invalid file type specified'}), 400

    if not session.get(file_path_session_key):
        return jsonify({'error': f'{file_type} not uploaded or session expired'}), 400
    if not sheet_name:
        return jsonify({'error': 'Sheet name not provided'}), 400

    try:
        df = pd.read_excel(session[file_path_session_key], sheet_name=sheet_name, nrows=0)
        return jsonify({'columns': df.columns.tolist()})
    except Exception as e:
        return jsonify({'error': f'Error reading sheet columns: {str(e)}'}), 500

@app.route('/configure', methods=['GET', 'POST'])
@login_required # Protect this route with login
def configure():
    if not session.get('db_uploaded'):
        flash('Please upload the main database first.', 'warning')
        return redirect(url_for('admin'))

    if request.method == 'POST':
        try:
            selected_sheet = request.form.get('sheet_name')
            key_field = request.form.get('key_field')

            if not selected_sheet or not key_field:
                raise ValueError("Please select both a sheet and a key field.")

            session['admin_selected_sheet'] = selected_sheet
            session['admin_key_field'] = key_field
            
            db_df_sample = pd.read_excel(session['db_file_path'], sheet_name=selected_sheet, nrows=0)
            session['admin_sheet_columns'] = db_df_sample.columns.tolist()

            session['db_configured'] = True
            flash('Admin configuration saved successfully!', 'success')
            return redirect(url_for('user_input'))
        except Exception as e:
            flash(f'Error saving configuration: {str(e)}', 'error')
            return redirect(url_for('configure'))

    sheet_names = session.get('db_sheet_names', [])
    return render_template('configure.html', sheet_names=sheet_names)


@app.route('/user_input', methods=['GET', 'POST'])
def user_input():
    # User input page doesn't strictly need login, but if it relies on admin config,
    # the admin config check will handle it.
    if not session.get('db_configured'):
        flash('Please upload and configure the main database first by an Admin.', 'warning')
        # Redirect to login if not configured, allowing admin to login and configure
        return redirect(url_for('login')) 

    if request.method == 'POST':
        if 'userfile' not in request.files:
            flash('No file part selected for user input', 'error')
            return redirect(request.url)
        file = request.files['userfile']
        if file.filename == '':
            flash('No user file selected', 'error')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            try:
                filename = secure_filename('user_input_file' + os.path.splitext(file.filename)[1])
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                xls = pd.ExcelFile(filepath)
                sheet_names = xls.sheet_names

                session['user_file_path'] = filepath
                session['user_sheet_names'] = sheet_names
                session['user_input_uploaded'] = True
                session.pop('user_input_configured', None)
                session.pop('processed_results_path', None)
                session.pop('result_columns', None)
                session.pop('processing_done', None)

                flash('User file uploaded. Please configure your input.', 'success')
                return redirect(url_for('user_configure'))
            except Exception as e:
                flash(f'Error processing user file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Invalid file type for user input. Allowed types are xlsx, xls, csv.', 'error')
            return redirect(request.url)
            
    return render_template('user_input.html')

@app.route('/user_configure', methods=['GET', 'POST'])
def user_configure():
    if not session.get('user_input_uploaded'):
        flash('Please upload your input file first.', 'warning')
        return redirect(url_for('user_input'))
    
    admin_key_field = session.get('admin_key_field', 'N/A')

    if request.method == 'POST':
        try:
            selected_sheet = request.form.get('sheet_name')
            parameters_to_compare = request.form.getlist('parameters')

            if not selected_sheet:
                raise ValueError("Please select a sheet from your file.")
            if not parameters_to_compare:
                raise ValueError("Please select at least one parameter for comparison.")

            session['user_selected_sheet'] = selected_sheet
            session['user_parameters_to_compare'] = parameters_to_compare
            
            user_df_sample = pd.read_excel(session['user_file_path'], sheet_name=selected_sheet, nrows=0)
            session['user_sheet_columns'] = user_df_sample.columns.tolist()

            session['user_input_configured'] = True
            
            session.pop('processed_results_path', None)
            session.pop('result_columns', None)
            session.pop('processing_done', None)
            
            flash('User configuration saved. Processing data...', 'success')
            return redirect(url_for('process_data'))
        except Exception as e:
            flash(f'Error saving user configuration: {str(e)}', 'error')
            return redirect(url_for('user_configure'))

    sheet_names = session.get('user_sheet_names', [])
    return render_template('user_configure.html', sheet_names=sheet_names, admin_key_field=admin_key_field)

@app.route('/process_data', methods=['GET'])
def process_data():
    if not session.get('user_input_configured'):
        flash('User input not configured. Please complete the configuration steps.', 'danger')
        return redirect(url_for('user_configure'))
    if not session.get('db_configured'):
        flash('Admin database not configured. Please complete admin steps.', 'danger')
        return redirect(url_for('configure'))

    try:
        db_df = pd.read_excel(session['db_file_path'], sheet_name=session['admin_selected_sheet'])
        user_df = pd.read_excel(session['user_file_path'], sheet_name=session['user_selected_sheet'])

        admin_key_field = session['admin_key_field']
        parameters_to_compare = session['user_parameters_to_compare']

        if admin_key_field not in db_df.columns:
            flash(f"Admin key field '{admin_key_field}' not found in the configured database sheet.", "danger")
            return redirect(url_for('configure'))
        
        if admin_key_field not in user_df.columns:
            flash(f"Admin key field '{admin_key_field}' not found in your uploaded user file sheet.", "danger")
            return redirect(url_for('user_configure'))
        
        user_ref_col_for_matching = admin_key_field

        filtered_results_list_of_dicts = [] 
        db_ref_values = db_df[admin_key_field].tolist()

        for _, user_row in user_df.iterrows():
            user_ref_val = user_row[user_ref_col_for_matching]
            db_match_key_val = advanced_matcher(user_ref_val, db_ref_values)

            filtered_result_row_dict = {
                'Input_Key_Value': user_ref_val,
                'Reference_Matched_Key_Value': db_match_key_val if db_match_key_val else 'Not Found'
            }
            
            matched_db_row = None
            if db_match_key_val and db_match_key_val != 'Not Found':
                matched_db_rows = db_df[db_df[admin_key_field].astype(str).str.strip().str.lower() == str(db_match_key_val).strip().lower()]
                if not matched_db_rows.empty:
                    matched_db_row = matched_db_rows.iloc[0]

            for param in parameters_to_compare:
                user_val = user_row.get(param, None)
                db_val = None
                if matched_db_row is not None:
                    db_val = matched_db_row.get(param, None)
                
                filtered_result_row_dict[f'Comparison_Result_for_{param}'] = compare_values(user_val, db_val)

            filtered_results_list_of_dicts.append(filtered_result_row_dict)

        if filtered_results_list_of_dicts:
            comparison_results_df = pd.DataFrame(filtered_results_list_of_dicts)
            
            unique_filename = f"processed_results_{uuid.uuid4().hex}.parquet"
            processed_filepath = os.path.join(app.config['PROCESSED_FOLDER'], unique_filename)
            
            comparison_results_df.to_parquet(processed_filepath, index=False)
            
            session['processed_results_path'] = processed_filepath
            session['result_columns'] = list(comparison_results_df.columns)
            session['processing_done'] = True
            flash('Data processing complete. Results are ready!', 'success')
        else:
            session['processed_results_path'] = None
            session['result_columns'] = ['Input_Key_Value', 'Reference_Matched_Key_Value'] 
            session['processing_done'] = True
            flash('Processing completed, but no matches were found between your file and the database for the selected key field.', 'warning')
        
        return redirect(url_for('results'))
        
    except Exception as e:
        app.logger.error(f"An error occurred during data processing: {str(e)}", exc_info=True)
        flash(f'An unexpected error occurred during data processing: {str(e)}. Please check your file formats, selected sheets, or admin configurations.', 'danger')
        
        session.pop('processed_results_path', None)
        session.pop('result_columns', None)
        session.pop('processing_done', None)
        return redirect(url_for('user_configure'))


@app.route('/results', methods=['GET'])
def results():
    if not session.get('processing_done', False):
        flash('No results available. Please ensure files are uploaded and processed.', 'error')
        return redirect(url_for('user_input'))

    processed_filepath = session.get('processed_results_path')
    titles = session.get('result_columns', [])

    results_data = []
    total_rows = 0

    if processed_filepath and os.path.exists(processed_filepath):
        try:
            full_results_df = pd.read_parquet(processed_filepath)
            results_data = full_results_df.to_dict(orient='records')
            total_rows = len(results_data)
        except Exception as e:
            app.logger.error(f"Error loading processed data from {processed_filepath}: {str(e)}", exc_info=True)
            flash('Error loading results data. The processed file might be corrupted or missing.', 'danger')
            session.pop('processed_results_path', None)
            return redirect(url_for('user_input'))
    else:
        if session.get('processing_done', False) and not processed_filepath:
             flash('Processing completed, but no results were generated.', 'info')

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    if per_page == 0: 
        per_page_effective = total_rows if total_rows > 0 else 20
    else:
        per_page_effective = per_page

    start = (page - 1) * per_page_effective
    end = start + per_page_effective
    paginated_results = results_data[start:end]

    prev_url = url_for('results', page=page - 1, per_page=per_page) if page > 1 else None
    next_url = url_for('results', page=page + 1, per_page=per_page) if end < total_rows else None
    
    all_rows_val = total_rows if total_rows > 0 else 20 

    return render_template('results.html',
                           results=paginated_results,
                           titles=titles,
                           page=page,
                           per_page=per_page,
                           total_rows=total_rows,
                           all_rows_val=all_rows_val,
                           prev_url=prev_url,
                           next_url=next_url)

@app.route('/download')
def download():
    processed_filepath = session.get('processed_results_path')
    if not processed_filepath or not os.path.exists(processed_filepath):
        flash('No results to download. Please process files first.', 'error')
        return redirect(url_for('user_input'))

    try:
        results_df = pd.read_parquet(processed_filepath)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Matching Results"

        for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True)):
            ws.append(row)
            if r_idx == 0:
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
        
        for col_idx, column_title in enumerate(results_df.columns):
            column_letter = get_column_letter(col_idx + 1)
            max_length = len(column_title)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        output_filename = 'matching_output.xlsx'
        output_path = os.path.join('static', output_filename)
        wb.save(output_path)

        response = make_response(send_from_directory('static', output_filename, as_attachment=True))
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response
    except Exception as e:
        app.logger.error(f"Error generating download file: {str(e)}", exc_info=True)
        flash(f"Error generating download file: {str(e)}", "danger")
        return redirect(url_for('results'))


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')